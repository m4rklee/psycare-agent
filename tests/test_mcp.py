import logging
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor

import pytest
from openpyxl import load_workbook

from app.api import mcp as mcp_api
from app.core.config import Settings
from app.services.tools import ExcelReportWriter


@pytest.mark.asyncio
async def test_mcp_initialize_and_tools_list() -> None:
    init = await mcp_api.mcp({"jsonrpc": "2.0", "id": 1, "method": "initialize", "params": {}})
    assert init["jsonrpc"] == "2.0"
    assert init["result"]["protocolVersion"] == "2024-11-05"

    tools = await mcp_api.mcp({"jsonrpc": "2.0", "id": 2, "method": "tools/list", "params": {}})
    names = {tool["name"] for tool in tools["result"]["tools"]}
    assert "multimodalAgent.excel.write_report" in names
    assert "multimodalAgent.email.send_alert" in names


@pytest.mark.asyncio
async def test_mcp_unknown_method_returns_json_rpc_error() -> None:
    response = await mcp_api.mcp({"jsonrpc": "2.0", "id": "x", "method": "missing", "params": {}})
    assert response["error"]["code"] == -32601


def test_excel_writer_payload_creates_workbook(tmp_path: Path) -> None:
    path = tmp_path / "reports.xlsx"
    writer = ExcelReportWriter(Settings(mcp_excel_local_path=str(path), ai_provider="mock"))
    writer.write_payload(
        {
            "reportId": 1,
            "userId": 2,
            "username": "student",
            "sessionId": "abc",
            "intent": "RISK",
            "emotion": "HIGH_RISK",
            "emotionScore": 4.0,
            "riskLevel": "HIGH",
            "confidence": 0.95,
            "summary": "summary",
            "emotionTags": None,
            "content": "content",
            "createdAt": "2026-06-08T00:00:00",
        }
    )
    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.cell(1, 1).value == "报告ID"
    assert sheet.cell(2, 3).value == "student"


def test_excel_writer_payload_serializes_multiple_writes(tmp_path: Path) -> None:
    path = tmp_path / "reports.xlsx"
    writer = ExcelReportWriter(Settings(mcp_excel_local_path=str(path), ai_provider="mock"))

    def write_row(index: int) -> None:
        writer.write_payload(
            {
                "reportId": index,
                "userId": 2,
                "username": f"student-{index}",
                "sessionId": "abc",
                "intent": "RISK",
                "emotion": "HIGH_RISK",
                "emotionScore": 4.0,
                "riskLevel": "HIGH",
                "confidence": 0.95,
                "summary": "summary",
                "emotionTags": None,
                "content": "content",
                "createdAt": "2026-06-08T00:00:00",
            }
        )

    with ThreadPoolExecutor(max_workers=4) as executor:
        list(executor.map(write_row, range(1, 9)))

    workbook = load_workbook(path)
    sheet = workbook.active
    assert sheet.max_row == 9
    assert sheet.cell(1, 1).value == "报告ID"
    assert {sheet.cell(row, 3).value for row in range(2, 10)} == {
        f"student-{index}" for index in range(1, 9)
    }


@pytest.mark.asyncio
async def test_mcp_email_tool_logs_alert(caplog: pytest.LogCaptureFixture) -> None:
    caplog.set_level(logging.WARNING, logger="app.services.tools")
    response = await mcp_api.mcp(
        {
            "jsonrpc": "2.0",
            "id": 3,
            "method": "tools/call",
            "params": {
                "name": "multimodalAgent.email.send_alert",
                "arguments": {
                    "recipient": "counselor@example.com",
                    "reportId": 42,
                    "username": "student",
                    "riskLevel": "HIGH",
                    "summary": "summary",
                },
            },
        }
    )
    assert response["result"]["content"][0]["text"] == "High-risk alert recorded through MCP protocol."
    assert "MCP high-risk alert" in caplog.text
    assert "recipient=counselor@example.com" in caplog.text
