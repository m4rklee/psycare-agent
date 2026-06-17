import logging
import smtplib
import threading
from collections.abc import Mapping
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

import httpx
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.config import Settings
from app.models.entities import AlertRecord, PsychologicalReport
from app.models.enums import RiskLevel, ToolStatus

logger = logging.getLogger(__name__)


class ExcelReportWriter:
    _write_lock = threading.RLock()

    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    async def write(self, report: PsychologicalReport) -> None:
        if self.settings.mcp_excel_mode == "http":
            await self._write_http(report)
            return
        if self.settings.mcp_excel_mode == "mcp":
            try:
                await self._write_mcp(report)
                return
            except Exception:
                pass
        self._write_local(report)

    async def _write_http(self, report: PsychologicalReport) -> None:
        async with httpx.AsyncClient(timeout=10) as client:
            response = await client.post(self.settings.mcp_excel_url, json=self._payload(report))
            response.raise_for_status()

    async def _write_mcp(self, report: PsychologicalReport) -> None:
        client = McpProtocolClient(self.settings.mcp_excel_url)
        await client.call_tool("multimodalAgent.excel.write_report", self._payload(report))

    def _write_local(self, report: PsychologicalReport) -> None:
        self.write_payload(self._payload(report))

    def write_payload(self, payload: Mapping[str, Any]) -> None:
        with self._write_lock:
            path = Path(self.settings.mcp_excel_local_path)
            path.parent.mkdir(parents=True, exist_ok=True)
            headers = [
                "报告ID",
                "用户ID",
                "账号",
                "会话ID",
                "意图",
                "情绪标签",
                "情绪总分",
                "风险等级",
                "置信度",
                "判断摘要",
                "多模态标签",
                "对话内容",
                "对话时间",
            ]
            if path.exists():
                workbook = load_workbook(path)
                sheet = workbook.active
            else:
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(headers)
            sheet.append(
                [
                    payload.get("reportId", 0),
                    payload.get("userId", 0),
                    payload.get("username", ""),
                    payload.get("sessionId", ""),
                    payload.get("intent", ""),
                    payload.get("emotion", ""),
                    payload.get("emotionScore", 0.0),
                    payload.get("riskLevel", ""),
                    payload.get("confidence", 0.0),
                    payload.get("summary", ""),
                    payload.get("emotionTags", ""),
                    payload.get("content", ""),
                    payload.get("createdAt", ""),
                ]
            )
            workbook.save(path)
            workbook.close()

    def _payload(self, report: PsychologicalReport) -> dict:
        return {
            "reportId": report.id,
            "userId": report.user.id,
            "username": report.user.username,
            "sessionId": report.session.public_id if report.session else None,
            "intent": report.intent.value,
            "emotion": report.emotion.value,
            "emotionScore": report.emotion_score,
            "riskLevel": report.risk_level.value,
            "confidence": report.confidence,
            "summary": report.summary,
            "emotionTags": report.emotion_tags,
            "content": report.content,
            "createdAt": report.created_at.isoformat() if report.created_at else None,
        }


class AlertNotifier:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    async def notify(self, alert: AlertRecord, report: PsychologicalReport) -> None:
        if self.settings.mcp_email_mode == "log":
            self.log_alert(self._payload(alert, report))
            return
        if self.settings.mcp_email_mode == "http":
            await self._notify_http(alert, report)
            return
        if self.settings.mcp_email_mode == "mcp":
            try:
                await self._notify_mcp(alert, report)
                return
            except Exception:
                pass
        self._notify_smtp(alert, report)

    async def _notify_http(self, alert: AlertRecord, report: PsychologicalReport) -> None:
        async with httpx.AsyncClient(timeout=10) as client:
            response = await client.post(self.settings.mcp_email_url, json=self._payload(alert, report))
            response.raise_for_status()

    async def _notify_mcp(self, alert: AlertRecord, report: PsychologicalReport) -> None:
        client = McpProtocolClient(self.settings.mcp_email_url)
        await client.call_tool("multimodalAgent.email.send_alert", self._payload(alert, report))

    def _notify_smtp(self, alert: AlertRecord, report: PsychologicalReport) -> None:
        message = EmailMessage()
        message["From"] = self.settings.alert_mail_from
        message["To"] = alert.recipient
        message["Subject"] = f"multimodalAgent 高风险预警 #{report.id}"
        message.set_content(
            f"学生：{report.user.username}\n风险：{report.risk_level.value}\n摘要：{report.summary}\n\n内容：{report.content}"
        )
        smtp_cls = smtplib.SMTP_SSL if self.settings.mail_smtp_ssl_enable else smtplib.SMTP
        with smtp_cls(self.settings.mail_host, self.settings.mail_port, timeout=10) as smtp:
            if self.settings.mail_smtp_starttls_enable:
                smtp.starttls()
            if self.settings.mail_smtp_auth and self.settings.mail_username:
                smtp.login(self.settings.mail_username, self.settings.mail_password)
            smtp.send_message(message)

    def _payload(self, alert: AlertRecord, report: PsychologicalReport) -> dict:
        return {
            "alertId": alert.id,
            "reportId": report.id,
            "recipient": alert.recipient,
            "username": report.user.username,
            "riskLevel": report.risk_level.value,
            "summary": report.summary,
            "content": report.content,
        }

    @staticmethod
    def log_alert(payload: Mapping[str, Any]) -> None:
        logger.warning(
            "MCP high-risk alert: "
            "recipient=%s, reportId=%s, user=%s, risk=%s, summary=%s",
            payload.get("recipient", ""),
            payload.get("reportId", ""),
            payload.get("username", ""),
            payload.get("riskLevel", ""),
            payload.get("summary", ""),
        )


class McpProtocolClient:
    def __init__(self, endpoint: str) -> None:
        self.endpoint = endpoint
        self.next_id = 1
        self.initialized = False

    async def call_tool(self, name: str, arguments: Mapping[str, Any]) -> Any:
        await self._ensure_initialized()
        await self._request("tools/list", {})
        return await self._request("tools/call", {"name": name, "arguments": dict(arguments)})

    async def _ensure_initialized(self) -> None:
        if self.initialized:
            return
        await self._request(
            "initialize",
            {
                "protocolVersion": "2024-11-05",
                "clientInfo": {"name": "multimodalAgent-agent", "version": "0.1.0"},
                "capabilities": {},
            },
        )
        self.initialized = True

    async def _request(self, method: str, params: Mapping[str, Any]) -> Any:
        request_id = self.next_id
        self.next_id += 1
        body = {"jsonrpc": "2.0", "id": request_id, "method": method, "params": dict(params)}
        async with httpx.AsyncClient(timeout=10) as client:
            response = await client.post(self.endpoint, json=body)
            response.raise_for_status()
            payload = response.json()
        if payload.get("error") is not None:
            raise RuntimeError(f"MCP error: {payload['error']}")
        return payload.get("result")


class ToolOrchestrationService:
    def __init__(
        self,
        settings: Settings,
        excel_writer: ExcelReportWriter,
        alert_notifier: AlertNotifier,
    ) -> None:
        self.settings = settings
        self.excel_writer = excel_writer
        self.alert_notifier = alert_notifier

    async def handle(self, session: AsyncSession, report_id: int) -> None:
        report = await session.scalar(
            select(PsychologicalReport)
            .options(
                selectinload(PsychologicalReport.user),
                selectinload(PsychologicalReport.session),
            )
            .where(PsychologicalReport.id == report_id)
        )
        if not report:
            return
        try:
            await self.excel_writer.write(report)
            report.excel_status = ToolStatus.SUCCESS
        except Exception as exc:
            report.excel_status = ToolStatus.FAILED
            report.tool_error = str(exc)[:500]
        if report.risk_level == RiskLevel.HIGH and report.excel_status == ToolStatus.SUCCESS:
            await self._send_alerts(session, report)
        await session.commit()

    async def _send_alerts(self, session: AsyncSession, report: PsychologicalReport) -> None:
        all_success = True
        for recipient in self.settings.alert_recipients:
            alert = AlertRecord(report=report, recipient=recipient)
            session.add(alert)
            await session.flush()
            sent = False
            for _ in range(max(1, self.settings.alert_mail_max_retries + 1)):
                alert.attempts += 1
                alert.updated_at = datetime.utcnow()
                try:
                    await self.alert_notifier.notify(alert, report)
                    alert.status = ToolStatus.SUCCESS
                    sent = True
                    break
                except Exception as exc:
                    alert.status = ToolStatus.FAILED
                    alert.error_message = str(exc)[:500]
            all_success = all_success and sent
        report.email_status = ToolStatus.SUCCESS if all_success else ToolStatus.FAILED
